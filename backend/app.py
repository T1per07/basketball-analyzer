"""FastAPI 主应用 - 篮球投篮分析 API"""
import uuid
import asyncio
import json
import base64
import io
import logging
import time
from pathlib import Path
from dataclasses import dataclass

import cv2
import numpy as np
from fastapi import FastAPI, UploadFile, File, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

from services.video_processor import VideoProcessor
from services.stats_calculator import StatsCalculator
from services.realtime_service import RealtimeAnalyzer
from config.settings import config
from utils.logger import performance_monitor

logger = logging.getLogger(__name__)

# ── 安全配置（从 config 统一读取）──
MAX_UPLOAD_SIZE_BYTES = config.security.max_upload_size_mb * 1024 * 1024
ALLOWED_VIDEO_EXTENSIONS = set(config.security.allowed_video_extensions)
ALLOWED_ORIGINS = list(config.security.allowed_origins)

app = FastAPI(
    title="Basketball Shot Analyzer",
    description="篮球投篮分析系统 - 检测、跟踪、轨迹分析、统计",
    version="0.1.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 存储分析任务状态
@dataclass
class TaskState:
    task_id: str
    status: str  # "processing", "completed", "failed"
    progress: float
    video_path: str
    result: dict | None = None
    error: str | None = None


tasks: dict[str, TaskState] = {}
_video_tasks: set[asyncio.Task] = set()  # 持有异步任务引用，防止 GC 回收
video_processor = VideoProcessor()
stats_calculator = StatsCalculator()


@app.get("/api")
async def api_info():
    return {
        "name": "Basketball Shot Analyzer",
        "version": "0.1.0",
        "endpoints": {
            "upload": "POST /api/v1/analyze",
            "status": "GET /api/v1/status/{task_id}",
            "results": "GET /api/v1/results/{task_id}",
            "video": "GET /api/v1/videos/{task_id}",
        },
    }


@app.post("/api/v1/analyze")
async def analyze_video(file: UploadFile = File(...)):
    """上传视频并开始分析"""
    # 格式白名单校验
    if not file.filename:
        raise HTTPException(400, "Filename is required")
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_VIDEO_EXTENSIONS:
        raise HTTPException(400, f"Unsupported video format: {ext}. Allowed: {', '.join(ALLOWED_VIDEO_EXTENSIONS)}")

    if not file.content_type or not file.content_type.startswith("video/"):
        raise HTTPException(400, "File must be a video")

    task_id = str(uuid.uuid4())[:8]
    upload_dir = config.video_upload_dir
    upload_dir.mkdir(parents=True, exist_ok=True)

    # 流式读取并限制大小，防止超大文件耗尽内存
    video_path = str(upload_dir / f"{task_id}_{file.filename}")
    total_size = 0
    chunk_size = 1024 * 1024  # 1MB per chunk
    with open(video_path, "wb") as f:
        while True:
            chunk = await file.read(chunk_size)
            if not chunk:
                break
            total_size += len(chunk)
            if total_size > MAX_UPLOAD_SIZE_BYTES:
                f.close()
                Path(video_path).unlink(missing_ok=True)
                raise HTTPException(413, f"File too large. Max size: {MAX_UPLOAD_SIZE_BYTES // (1024*1024)}MB")
            f.write(chunk)

    tasks[task_id] = TaskState(
        task_id=task_id,
        status="processing",
        progress=0.0,
        video_path=video_path,
    )

    # 异步处理视频（持有任务引用防止 GC 回收）
    task = asyncio.create_task(_process_video(task_id, video_path))
    _video_tasks.add(task)
    task.add_done_callback(_video_tasks.discard)

    return {
        "task_id": task_id,
        "status": "processing",
        "message": "Video uploaded, analysis started",
    }


async def _process_video(task_id: str, video_path: str):
    """异步处理视频（在后台线程中运行，不阻塞事件循环）"""
    try:
        def progress_cb(current, total):
            tasks[task_id].progress = current / total if total > 0 else 0

        def run_analysis():
            result = video_processor.analyze_video(video_path, progress_cb)
            # 获取视频真实分辨率，用于热力图坐标归一化
            video_info = video_processor.get_video_info(video_path)
            full_stats = stats_calculator.generate_full_stats(
                result,
                image_width=video_info.width,
                image_height=video_info.height,
            )

            # 添加投篮详情（含运动学参数）
            full_stats["shots"] = [
                {
                    "shot_id": int(s.shot_id),
                    "shot_type": str(s.shot_type),
                    "made": bool(s.made),
                    "distance": float(s.distance),
                    "release_angle": float(s.release_angle),
                    "entry_angle": float(s.entry_angle),
                    "confidence": float(s.confidence),
                    "frame": int(s.frame_end),
                    "flight_time": float(s.flight_time),
                    "shot_speed": float(s.shot_speed),
                    "arc_height": float(s.arc_height),
                }
                for s in result.shots
            ]

            # 运动学汇总
            full_stats["kinematics"] = {
                "avg_speed": round(result.average_speed, 2),
                "avg_flight_time": round(result.average_flight_time, 3),
                "avg_arc_height": round(result.average_arc_height, 2),
            }

            # 生成标注视频（60fps）
            output_dir = config.output_dir
            output_dir.mkdir(parents=True, exist_ok=True)
            annotated_path = str(output_dir / f"{task_id}_annotated.mp4")

            video_processor.create_annotated_video(
                video_path, annotated_path, result, progress_cb
            )

            return full_stats

        full_stats = await asyncio.to_thread(run_analysis)

        tasks[task_id].status = "completed"
        tasks[task_id].progress = 1.0
        tasks[task_id].result = {
            **full_stats,
            "annotated_video": f"/api/v1/videos/{task_id}",
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        tasks[task_id].status = "failed"
        tasks[task_id].error = str(e)


@app.get("/api/v1/status/{task_id}")
async def get_status(task_id: str):
    """获取分析进度"""
    if task_id not in tasks:
        raise HTTPException(404, "Task not found")

    task = tasks[task_id]
    return {
        "task_id": task.task_id,
        "status": task.status,
        "progress": task.progress,
    }


@app.get("/api/v1/results/{task_id}")
async def get_results(task_id: str):
    """获取分析结果"""
    if task_id not in tasks:
        raise HTTPException(404, "Task not found")

    task = tasks[task_id]
    if task.status == "processing":
        raise HTTPException(202, "Analysis still in progress")
    if task.status == "failed":
        raise HTTPException(500, f"Analysis failed: {task.error}")

    return {
        "task_id": task.task_id,
        "status": task.status,
        **task.result,
    }


@app.get("/api/v1/videos/{task_id}")
async def get_video(task_id: str):
    """获取标注后的视频"""
    if task_id not in tasks:
        raise HTTPException(404, "Task not found")

    video_path = config.output_dir / f"{task_id}_annotated.mp4"
    if not video_path.exists():
        raise HTTPException(404, "Video not found")

    return FileResponse(str(video_path), media_type="video/mp4")


@app.get("/api/v1/health")
async def health():
    return {"status": "ok"}


@app.get("/api/v1/metrics")
async def metrics():
    """获取性能监控指标"""
    return {
        "status": "ok",
        "metrics": performance_monitor.get_all_stats(),
        "active_tasks": len([t for t in tasks.values() if t.status == "processing"]),
        "total_tasks": len(tasks),
    }


@app.get("/api/v1/export/excel/{task_id}")
async def export_excel(task_id: str):
    """导出 Excel 报告"""
    if task_id not in tasks or tasks[task_id].status != "completed":
        raise HTTPException(404, "Result not found")

    result = tasks[task_id].result
    shots = result.get("shots", [])
    summary = result.get("summary", {})
    by_type = result.get("by_type", {})
    kinematics = result.get("kinematics", {})

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    wb = Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="FF6B2B", end_color="FF6B2B", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    ws.append(["Basketball Shot Analysis Report"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    for label, key in [("Total Shots", "total_shots"), ("Made", "made_shots"),
                       ("FG%", "overall_percentage"), ("Avg Distance (m)", "average_distance"),
                       ("Video Duration (s)", "video_duration")]:
        val = summary.get(key, 0)
        if key == "overall_percentage":
            val = f"{val:.1%}"
        elif isinstance(val, float):
            val = round(val, 2)
        ws.append([label, val])

    if kinematics:
        ws.append([])
        ws.append(["Kinematics"])
        ws["A" + str(ws.max_row)].font = header_font
        for label, key in [("Avg Speed (m/s)", "avg_speed"),
                           ("Avg Flight Time (s)", "avg_flight_time"),
                           ("Avg Arc Height (m)", "avg_arc_height")]:
            ws.append([label, kinematics.get(key, 0)])

    # By Type sheet
    ws2 = wb.create_sheet("By Type")
    ws2.append(["Type", "Attempts", "Made", "FG%", "Avg Distance"])
    for cell in ws2[1]:
        cell.font = header_font_white
        cell.fill = header_fill
    for t, s in by_type.items():
        ws2.append([t, s["attempts"], s["made"], f"{s['percentage']:.1%}", round(s["avg_distance"], 2)])

    # Shots sheet
    ws3 = wb.create_sheet("All Shots")
    headers = ["ID", "Type", "Made", "Distance (m)", "Release Angle", "Entry Angle",
               "Speed (m/s)", "Flight Time (s)", "Arc Height (m)", "Confidence"]
    ws3.append(headers)
    for cell in ws3[1]:
        cell.font = header_font_white
        cell.fill = header_fill
    for s in shots:
        ws3.append([
            s["shot_id"], s["shot_type"], "Yes" if s["made"] else "No",
            round(s["distance"], 2), round(s["release_angle"], 1),
            round(s["entry_angle"], 1), round(s.get("shot_speed", 0), 2),
            round(s.get("flight_time", 0), 3), round(s.get("arc_height", 0), 2),
            round(s["confidence"], 2),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=shot_analysis_{task_id}.xlsx"},
    )


@app.get("/api/v1/export/pdf/{task_id}")
async def export_pdf(task_id: str):
    """导出 PDF 报告"""
    if task_id not in tasks or tasks[task_id].status != "completed":
        raise HTTPException(404, "Result not found")

    result = tasks[task_id].result
    shots = result.get("shots", [])
    summary = result.get("summary", {})
    by_type = result.get("by_type", {})
    kinematics = result.get("kinematics", {})

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        raise HTTPException(500, "reportlab not installed. Run: pip install reportlab")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title2", parent=styles["Title"], fontSize=18, spaceAfter=12)
    subtitle_style = ParagraphStyle("Sub", parent=styles["Heading2"], fontSize=13, spaceBefore=8, spaceAfter=6)

    elements = []

    elements.append(Paragraph("Basketball Shot Analysis Report", title_style))
    elements.append(Spacer(1, 6 * mm))

    # Summary table
    elements.append(Paragraph("Summary", subtitle_style))
    summary_data = [
        ["Metric", "Value"],
        ["Total Shots", str(summary.get("total_shots", 0))],
        ["Made", str(summary.get("made_shots", 0))],
        ["FG%", f"{summary.get('overall_percentage', 0):.1%}"],
        ["Avg Distance", f"{summary.get('average_distance', 0):.2f} m"],
    ]
    if kinematics:
        summary_data.extend([
            ["Avg Speed", f"{kinematics.get('avg_speed', 0):.2f} m/s"],
            ["Avg Flight Time", f"{kinematics.get('avg_flight_time', 0):.3f} s"],
            ["Avg Arc Height", f"{kinematics.get('avg_arc_height', 0):.2f} m"],
        ])

    t = Table(summary_data, colWidths=[120, 100])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF6B2B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 6 * mm))

    # By type
    if by_type:
        elements.append(Paragraph("Shot Types", subtitle_style))
        type_data = [["Type", "Attempts", "Made", "FG%", "Avg Dist (m)"]]
        for name, s in by_type.items():
            type_data.append([name, str(s["attempts"]), str(s["made"]),
                              f"{s['percentage']:.1%}", f"{s['avg_distance']:.2f}"])
        t2 = Table(type_data, colWidths=[80, 60, 50, 50, 80])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#333")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(t2)
        elements.append(Spacer(1, 6 * mm))

    # Shots table (first 50)
    elements.append(Paragraph("Shot Details (first 50)", subtitle_style))
    shot_data = [["#", "Type", "Made", "Dist", "Angle", "Speed", "Flight", "Arc"]]
    for s in shots[:50]:
        shot_data.append([
            str(s["shot_id"]), s["shot_type"], "Y" if s["made"] else "N",
            f"{s['distance']:.1f}", f"{s['release_angle']:.0f}°",
            f"{s.get('shot_speed', 0):.1f}", f"{s.get('flight_time', 0):.2f}s",
            f"{s.get('arc_height', 0):.1f}m",
        ])
    t3 = Table(shot_data, colWidths=[25, 60, 35, 40, 45, 45, 45, 45])
    t3.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f8f8")]),
    ]))
    elements.append(t3)

    doc.build(elements)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=shot_analysis_{task_id}.pdf"},
    )


@app.websocket("/ws/realtime")
async def websocket_realtime(websocket: WebSocket):
    """
    WebSocket 实时视频流分析
    客户端发送: base64 编码的 JPEG 帧
    服务端返回: 分析结果 + 标注后的帧
    """
    await websocket.accept()
    analyzer = RealtimeAnalyzer()

    try:
        while True:
            data = await websocket.receive_text()
            msg = json.loads(data)

            if msg.get("type") == "reset":
                analyzer.reset()
                await websocket.send_json({"type": "reset_ack"})
                continue

            if msg.get("type") != "frame":
                continue

            # 解码帧
            frame_data = base64.b64decode(msg["data"])
            np_arr = np.frombuffer(frame_data, np.uint8)
            frame = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)

            if frame is None:
                continue

            # 处理帧
            result = analyzer.process_frame(frame)

            # 编码标注帧
            annotated_b64 = analyzer.encode_frame(result["frame"])

            # 发送结果（不发送原始帧数据，节省带宽）
            await websocket.send_json({
                "type": "result",
                "frame": annotated_b64,
                "stats": result["stats"],
                "detections": result["detections"],
                "timestamp": time.time(),
            })

    except WebSocketDisconnect:
        pass
    except Exception as e:
        try:
            await websocket.send_json({"type": "error", "message": str(e)})
        except Exception:
            pass


@app.websocket("/ws/camera")
async def websocket_camera(websocket: WebSocket):
    """
    WebSocket 摄像头实时分析
    客户端发送: 控制指令 (start/stop/config)
    服务端返回: 实时分析结果流
    """
    await websocket.accept()
    analyzer = RealtimeAnalyzer()
    running = False
    cap = None

    try:
        while True:
            data = await websocket.receive_text()
            msg = json.loads(data)

            if msg.get("type") == "start":
                source = msg.get("source", 0)
                # 安全校验：只允许整数摄像头索引（0-10），禁止传入文件路径
                if isinstance(source, str):
                    try:
                        source = int(source)
                    except (ValueError, TypeError):
                        await websocket.send_json({"type": "error", "message": "Invalid camera source"})
                        continue
                if not isinstance(source, int) or source < 0 or source > 10:
                    await websocket.send_json({"type": "error", "message": "Camera source must be 0-10"})
                    continue
                cap = cv2.VideoCapture(source)
                if not cap.isOpened():
                    await websocket.send_json({"type": "error", "message": f"Cannot open camera: {source}"})
                    continue

                running = True
                analyzer.reset()
                await websocket.send_json({"type": "started", "source": source})

                # 逐帧读取并分析
                while running:
                    ret, frame = cap.read()
                    if not ret:
                        break

                    result = analyzer.process_frame(frame)
                    annotated_b64 = analyzer.encode_frame(result["frame"])

                    await websocket.send_json({
                        "type": "result",
                        "frame": annotated_b64,
                        "stats": result["stats"],
                        "detections": result["detections"],
                    })

                    # 检查是否有控制消息（非阻塞）
                    try:
                        ctrl = await asyncio.wait_for(
                            websocket.receive_text(), timeout=0.001
                        )
                        ctrl_msg = json.loads(ctrl)
                        if ctrl_msg.get("type") == "stop":
                            running = False
                            break
                    except asyncio.TimeoutError:
                        pass

                cap.release()
                cap = None
                await websocket.send_json({"type": "stopped"})

            elif msg.get("type") == "stop":
                running = False
                if cap:
                    cap.release()
                    cap = None

            elif msg.get("type") == "config":
                # 动态更新配置
                if "skip_frames" in msg:
                    analyzer.config.target_fps = msg["skip_frames"]
                if "jpeg_quality" in msg:
                    analyzer.config.jpeg_quality = msg["jpeg_quality"]
                await websocket.send_json({"type": "config_updated"})

    except WebSocketDisconnect:
        running = False
        if cap:
            cap.release()
    except Exception as e:
        running = False
        if cap:
            cap.release()
        try:
            await websocket.send_json({"type": "error", "message": str(e)})
        except Exception:
            pass


# Serve frontend static files (production mode)
frontend_dist = Path(__file__).parent.parent / "frontend" / "dist"
if frontend_dist.exists():
    app.mount("/assets", StaticFiles(directory=str(frontend_dist / "assets")), name="static-assets")

    @app.get("/{full_path:path}")
    async def serve_frontend(full_path: str):
        """Serve frontend files (SPA fallback)"""
        file_path = frontend_dist / full_path
        if file_path.is_file():
            return FileResponse(str(file_path))
        return FileResponse(str(frontend_dist / "index.html"))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
